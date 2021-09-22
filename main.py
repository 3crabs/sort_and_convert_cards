from threading import Thread
from tkinter import *
from tkinter import filedialog as fd
import functools
from openpyxl import load_workbook

template_file_path = "../resources/Шаблон.xlsx"
window = Tk()
file_path = StringVar()
dir_path = StringVar()
start_button = {}
message = StringVar()
message_label = None

rows = []


class Row:

    def __init__(self, amount, eng_name, rus_name, card_set, condition, lang, foil, price, color):
        self.amount = amount
        self.eng_name = eng_name
        self.rus_name = rus_name
        self.card_set = card_set
        self.condition = condition
        self.lang = lang
        self.foil = foil
        self.price = price
        self.color = color

    def __str__(self) -> str:
        return f'{self.amount} {self.eng_name} {self.rus_name} {self.card_set} ' \
               f'{self.condition} {self.lang} {self.foil} {self.price}'

    def format_output(self):
        rus_name = " "
        foil = " "

        if self.foil == 1:
            foil = " FOIL "

        if self.rus_name != "---":
            rus_name = f" / {self.rus_name} "
        return f'{self.amount} {self.eng_name}{rus_name}{self.card_set} ' \
               f'{self.condition} {self.lang}{foil}{self.price} \n'


def check_button_state():
    global start_button
    start_button["state"] = NORMAL
    if file_path.get() == '':
        start_button["state"] = DISABLED
    if dir_path.get() == '':
        start_button["state"] = DISABLED


def select_dir():
    global dir_path
    dir_name = fd.askdirectory(title='Выберите папку для сохранения результатов')
    dir_path.set(dir_name)
    check_button_state()


def select_file():
    global file_path
    file_name = fd.askopenfile(title='Выберите файл данных', filetypes=[('xlsx files', ['.xlsx'])])
    file_path.set(file_name.name)
    check_button_state()


def sort_cards(first_row, second_row):
    if first_row.card_set > second_row.card_set:
        return 1
    if first_row.card_set == second_row.card_set:
        if first_row.color > second_row.color:
            return 1
        elif first_row.color == second_row.color:
            if first_row.eng_name > second_row.eng_name:
                return 1
    return -1


def work():
    global message
    message.set('Чтение данных')
    workbook = load_workbook(file_path.get())
    message.set('Чтение данных завершено')
    ws = workbook.worksheets[0]
    i = 2
    while ws[i][0].value is not None:
        color = ws[i][4].value
        if color is None:
            color = 'Z'
        r = Row(ws[i][10].value, ws[i][0].value, ws[i][1].value,
                ws[i][3].value, ws[i][7].value, ws[i][6].value,
                ws[i][8].value, ws[i][11].value, color)
        rows.append(r)
        i += 1

    rows.sort(key=functools.cmp_to_key(sort_cards))

    with open(dir_path.get() + r'/cards.txt', 'w', encoding='utf-8') as txt_file:
        for i in range(len(rows)):
            txt_file.write(rows[i].format_output())
        message.set('Готово!')


def start():
    global file_path
    global dir_path
    global message

    start_button["state"] = DISABLED
    message.set('Обрабротка начата')
    thread = Thread(target=work)
    thread.start()


def main():
    global window
    global file_path
    global dir_path
    global start_button
    global message
    global message_label

    window.title("Нормально делаем")
    window.geometry('333x111')

    file_label = Label(text="Путь к файлу:")
    file_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
    file_input = Entry(textvariable=file_path)
    file_input.grid(row=0, column=1, padx=5, pady=5)
    file_button = Button(window, text="Выбрать файл", command=select_file)
    file_button.grid(column=2, row=0, padx=5, pady=5, sticky="e")

    dir_label = Label(text="Путь к папке:")
    dir_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
    dir_input = Entry(textvariable=dir_path)
    dir_input.grid(row=1, column=1, padx=5, pady=5)
    dir_button = Button(window, text="Выбрать папку", command=select_dir)
    dir_button.grid(column=2, row=1, padx=5, pady=5, sticky="e")

    message_label = Label(textvariable=message)
    message_label.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="w")
    message.set("Заполните поля, нажмите кнопку."[0:35])

    start_button = Button(window, text="Начать", command=start)
    start_button.grid(column=2, row=2, padx=5, pady=5, sticky="e")
    start_button["state"] = DISABLED

    window.mainloop()


if __name__ == '__main__':
    main()
